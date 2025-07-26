import pandas as pd
from django.core.files.storage import default_storage
from django.conf import settings
import unicodedata
import re
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class Relatorio:

    def __init__(self):
        self.caminho_relatorio = os.path.join(settings.MEDIA_ROOT, 'relatorio_atualizado.xlsx')
        self.dpt = os.path.join(settings.MEDIA_ROOT, 'DPTDIA.csv')
        if os.path.exists(self.dpt):
            os.remove(self.dpt)
        if os.path.exists(self.caminho_relatorio):
            os.remove(self.caminho_relatorio)


    def retorna(self, arquivo1, arquivo2):
        nome1 = self.limpar_nome(arquivo1.name)
        nome2 = self.limpar_nome(arquivo2.name)
        self.caminho1 = default_storage.save(nome1, arquivo1)
        self.caminho2 = default_storage.save(nome2, arquivo2)
        self.Converter()
        self.LinkPathPTD = os.path.join(settings.MEDIA_ROOT, 'DPTDIA.csv')
        self.departamento = pd.read_csv(self.LinkPathPTD, encoding='latin1')
        self.dias = self.Dias()
        self.Add()
        self.centros = self.CentroCustos()
        self.relatorio = pd.read_csv((os.path.join(settings.BASE_DIR, 'relatorio.csv')), encoding='latin1')

    def limpar_nome(self, nome):
        nome = unicodedata.normalize('NFKD', nome).encode('ASCII', 'ignore').decode('ASCII')
        nome = re.sub(r'[^\w\-_\. ]', '_', nome)
        return nome


    def Converter(self):
        self.caminho1 = default_storage.path(self.caminho1)
        self.caminho2 = default_storage.path(self.caminho2)
        relatorio1 = pd.read_excel(self.caminho1, parse_dates=["DATA"])
        relatorio2 = pd.read_excel(self.caminho2, parse_dates=["DATA"])
        relatorio1['DATA'] = relatorio1['DATA'].dt.strftime('%d/%m/%y')
        relatorio2['DATA'] = relatorio2['DATA'].dt.strftime('%d/%m/%y')
        self.relatorio = pd.concat([relatorio1, relatorio2], ignore_index=True)
        caminho_dptdia_csv = os.path.join(settings.MEDIA_ROOT, 'DPTDIA.csv')
        self.relatorio.to_csv(caminho_dptdia_csv, encoding='latin1', index=False)


    def Add(self):
        with open((os.path.join(settings.BASE_DIR, 'relatorio.csv')), 'r', encoding='latin1') as arq:
            arq = arq.readlines()
            with open((os.path.join(settings.BASE_DIR, 'relatorio.csv')), 'w', encoding='latin1') as csv:
                for i in arq:
                    csv.write(f'{i[:i.find(',') + 1]}{','.join(self.dias)},{i[i.find('Total'):]}')

    def remove(self, lista):
        return list(dict.fromkeys(sorted(lista)))

    def Dias(self):
        return self.remove(list(self.departamento['DATA']))

    def CentroCustos(self):
        centro = self.departamento['C.C']
        centro = self.remove(list(centro))
        lista = []
        for i in range(0, len(centro)):
            pre = str(centro[i])
            ic = f'{''.join((pre[0], pre[1]))}'
            cc = f'{''.join((pre[2], pre[3]))}'
            if ic == '81':
                emp = 'TVCA'
            if ic == '65':
                emp = 'Portal MT'
            if ic == '69':
                emp = 'FMCA'
            if ic == '85':
                emp = 'On Line(MT)'
            if cc == '03':
                tipo = 'ADM'
            if cc == '06':
                tipo = 'RH'
            if cc == '10':
                tipo = 'COMERCIA'
            if cc == '11':
                tipo = 'OPEC'
            if cc == '12':
                tipo = 'MKT'
            if cc == '14':
                tipo = 'PROGRAMAÇÃO'
            if cc == '15':
                tipo = 'JORNALISMO'
            if cc == '21':
                tipo = 'TECNOLOGIA'
            lista.append(f'{centro[i]} - {tipo} - {emp}')
        return self.remove(lista)

    def money(self):
        money = 0
        for cc in self.centros:
            for i in range(0, 1):
                nm = f'{''.join((cc[0], cc[1], cc[2], cc[3]))}'
                cc = int(nm)
            soma = 0
            for dia in self.dias:
                qtd = len(self.departamento[(self.departamento['C.C'] == cc) & (self.departamento['DATA'] == dia)])
                soma += qtd
            money += soma
        return money

    def ArrumarColunas(self):
        workbook = load_workbook(self.caminho_relatorio)
        sheet = workbook.active
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(self.caminho_relatorio)

    def Converte(self):
        linhas = []
        money = self.money()
        atual = {'DPT': 0}
        for dia in self.dias:
            total_dia = 0
            for cc in self.centros:
                for i in range(0, 1):
                    nm = f'{''.join((cc[0], cc[1], cc[2], cc[3]))}'
                    cc = int(nm)
                qtd = len(self.departamento[(self.departamento['C.C'] == cc) & (self.departamento['DATA'] == dia)])
                total_dia += qtd
            atual[dia] = total_dia
        total = 0
        for i in atual:
            total += atual[i]
        atual['Total'] = total
        atual['Valor total'] = f'{money * 20:.2f}'
        pct = []
        for cc in self.centros:
            nova_linha = {'DPT': cc}
            for i in range(0, 1):
                nm = f'{''.join((cc[0], cc[1], cc[2], cc[3]))}'
                cc = int(nm)
            soma = 0
            for dia in self.dias:
                qtd = len(self.departamento[(self.departamento['C.C'] == cc) & (self.departamento['DATA'] == dia)])
                nova_linha[dia] = qtd
                soma += qtd
            valor = float(f'{(soma * 20) / float(atual['Valor total'])}')
            nova_linha['%'] = f'{(valor * 100):.2f}%'
            pct.append(valor)
            nova_linha['Total'] = soma
            nova_linha['Valor total'] = f'R$ {float(soma * 20)}'
            self.relatorio = pd.concat([self.relatorio, pd.DataFrame([nova_linha])], ignore_index=True)

        atual['%'] = f'{round(sum(pct) * 100)}%'
        self.relatorio = pd.concat([self.relatorio, pd.DataFrame([atual])], ignore_index=True)
        self.relatorio = pd.concat([self.relatorio, pd.DataFrame([linhas])], ignore_index=True)
        self.relatorio.to_excel(os.path.join(settings.MEDIA_ROOT,'relatorio_atualizado.xlsx'), index=False)
        self.ArrumarColunas()
        caminho1_full = os.path.join(settings.MEDIA_ROOT, self.caminho1)
        if os.path.exists(caminho1_full):
            os.remove(caminho1_full)
        caminho2_full = os.path.join(settings.MEDIA_ROOT, self.caminho2)
        if os.path.exists(caminho2_full):
            os.remove(caminho2_full)
        nome_saida = 'relatorio_atualizado.xlsx'
        return f'{nome_saida}'